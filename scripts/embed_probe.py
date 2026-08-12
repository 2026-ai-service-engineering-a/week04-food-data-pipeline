"""임베딩 조립안 A/B 프로브 — 무엇을 벡터로 만들지 실험으로 정한다.

검색 품질은 모델이 아니라 **무엇을 임베딩하느냐**에서 갈리는데, 답은 데이터마다
다르다. 추측하지 말고 재보라는 것이 이 스크립트의 전부다.

이 데이터의 사정:
  - 식품명 평균 10.4자. `소금빵`, `스파게티`. 짧고 브랜드·외국어가 섞인다
  - 분류 체계는 결측 0%, 소분류만 219종. 사람이 붙여둔 라벨에 가깝다
  - 분류가 비면 '해당없음'이 들어가 있다. 30만 건에 공통으로 깔리는 문자열이다

조립안:
  A name            식품명만
  B name+cat_raw    식품명 + 대/중/소분류 (원문 그대로 — '해당없음' 포함)
  C name+cat        식품명 + 대/중/소분류 ('해당없음' 제거)     ← 기본 채택안
  D name+cat+maker  C + 제조사명

**작게 재고, 좁히고, 크게 확인한다.** 1,000건으로 네 조립안을 훑어 후보를 둘로
줄인 뒤, 전량으로 그 둘만 맞붙인다. 처음부터 전량 네 조립안을 돌리면 네 배 비싸고
네 배 오래 걸리는데, 얻는 결론은 같다.

사용법:
  # 1단계 — 샘플로 네 조립안 훑기 (1분, 2센트)
  docker compose exec api uv run python scripts/embed_probe.py

  # 2단계 — 전량으로 결승전 (약 30분, 2달러)
  docker compose exec api uv run python scripts/embed_probe.py \\
    --input data/raw/20260728_가공식품DB_306307건.xlsx --variants A,C --workers 8
"""

import argparse
import json
import os
import re
import threading
import time
from pathlib import Path

import numpy as np
import pandas as pd

MODEL = os.environ.get("EMBEDDING_MODEL", "gemini/gemini-embedding-001")
DIM = int(os.environ.get("EMBEDDING_DIM", "768"))
BATCH = int(os.environ.get("EMBED_BATCH_SIZE", "100"))
WORK = Path("data/clean")

NEEDED = ["식품코드", "식품명", "식품대분류명", "식품중분류명", "식품소분류명", "제조사명"]

# 분류가 비었을 때 원본이 쓰는 값. 30만 건에 공통으로 깔리는 문자열이다
EMPTY_CATEGORY = {"해당없음", "", "nan", "None"}

QUERIES = [
    "매콤한 분식 간식",
    "나트륨 낮은 튀김 간식",
    "아이 간식으로 좋은 부드러운 빵",
    "더울 때 시원하게 마실 음료",
    "밥 대신 간단히 먹는 즉석식품",
]

RETRY_HINT = re.compile(r"retry in ([\d.]+)s")


# ── 입력: csv · parquet · 187MB xlsx ──────────────────────────────────
def load_frame(path: str) -> pd.DataFrame:
    p = Path(path)
    if p.suffix == ".csv":
        return pd.read_csv(p, usecols=lambda c: c in NEEDED).fillna("")
    if p.suffix == ".parquet":
        return pd.read_parquet(p).fillna("")

    # 187MB xlsx는 로드만 몇 분이라 필요한 6열만 스트리밍으로 뽑아 캐시한다
    cache = WORK / ".probe_source.parquet"
    if cache.exists():
        print(f"[load] 캐시 사용 {cache}")
        return pd.read_parquet(cache).fillna("")

    from openpyxl import load_workbook

    print(f"[load] {p.name} 스트리밍 읽기 (필요한 {len(NEEDED)}열만)")
    ws = load_workbook(p, read_only=True, data_only=True)[
        load_workbook(p, read_only=True).sheetnames[0]
    ]
    rows = ws.iter_rows(values_only=True)
    header = [("" if c is None else str(c)) for c in next(rows)]
    idx = [header.index(c) for c in NEEDED]
    data = []
    for i, row in enumerate(rows, 1):
        data.append([("" if row[j] is None else str(row[j])) for j in idx])
        if i % 100000 == 0:
            print(f"       {i:,}행", flush=True)
    df = pd.DataFrame(data, columns=NEEDED)
    cache.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(cache, index=False)
    print(f"[load] {len(df):,}행 → 캐시 {cache}")
    return df


# ── 조립안 ────────────────────────────────────────────────────────────
def build_texts(df: pd.DataFrame, which: list[str]) -> dict[str, list[str]]:
    name = df["식품명"].astype(str).tolist()
    cats = [
        df["식품대분류명"].astype(str).tolist(),
        df["식품중분류명"].astype(str).tolist(),
        df["식품소분류명"].astype(str).tolist(),
    ]
    maker = df["제조사명"].astype(str).tolist()
    n = len(name)

    out = {}
    if "A" in which:
        out["A name"] = name
    if "B" in which:
        out["B name+cat_raw"] = [
            " · ".join([name[i]] + [c[i] for c in cats]) for i in range(n)
        ]
    if "C" in which or "D" in which:
        clean = [
            " · ".join([name[i]] + [c[i] for c in cats if c[i] not in EMPTY_CATEGORY])
            for i in range(n)
        ]
        if "C" in which:
            out["C name+cat"] = clean
        if "D" in which:
            out["D name+cat+maker"] = [
                clean[i] + (f" · {maker[i]}" if maker[i] not in EMPTY_CATEGORY else "")
                for i in range(n)
            ]
    return out


# ── 임베딩: 고유값으로 접고, 동시 호출하고, 재개한다 ──────────────────
def embed_unique(
    texts: list[str], task: str, tag: str, workers: int, stats: dict
) -> tuple[np.ndarray, np.ndarray]:
    """고유 텍스트만 임베딩하고 (행 → 고유값) 매핑을 함께 돌려준다.

    1회전에서 배운 것을 그대로 쓴다. 같은 문자열을 두 번 물어볼 이유가 없다.
    다만 여기서는 식품명이 대체로 고유해서 절감이 크지 않다 — 접히면 좋고
    아니면 안전장치로 간다.
    """
    import litellm

    uniq, inverse = np.unique(np.array(texts, dtype=object), return_inverse=True)
    uniq = uniq.tolist()
    n = len(uniq)
    vec_path = WORK / f".probe_{tag}.f32"
    done_path = WORK / f".probe_{tag}.done"
    WORK.mkdir(parents=True, exist_ok=True)

    done = 0
    if done_path.exists() and vec_path.exists():
        meta = json.loads(done_path.read_text())
        if meta.get("n") == n and meta.get("dim") == DIM:
            done = meta.get("done", 0)
            print(f"  [{tag}] 재개: {done:,}/{n:,} 이미 완료")
        else:
            vec_path.unlink(missing_ok=True)

    mm = np.memmap(vec_path, dtype=np.float32, mode="r+" if done else "w+", shape=(n, DIM))
    lock = threading.Lock()
    t0 = time.time()
    progress = {"n": done}

    def work(start: int):
        chunk = uniq[start : start + BATCH]
        for attempt in range(6):
            try:
                res = litellm.embedding(
                    model=MODEL, input=chunk, dimensions=DIM, task_type=task
                )
                break
            except Exception as exc:  # noqa: BLE001 — 429는 서버가 말한 만큼 기다린다
                if attempt == 5:
                    raise
                hint = RETRY_HINT.search(str(exc))
                wait = float(hint.group(1)) + 1 if hint else 2**attempt
                with lock:
                    stats["retries"] += 1
                    # 재시도는 보이게 남긴다. 조용히 삼키면 "429가 없었다"고
                    # 착각하고, 처리량이 왜 낮은지 영영 설명하지 못한다
                    if stats["retries"] % 100 == 0:
                        print(f"    429 누적 {stats['retries']:,}회 · "
                              f"{wait:.0f}s 대기", flush=True)
                time.sleep(wait)
        arr = np.array([d["embedding"] for d in res.data], dtype=np.float32)
        mm[start : start + len(chunk)] = arr
        with lock:
            stats["tokens"] += getattr(res.usage, "prompt_tokens", 0) or 0
            stats["calls"] += 1
            progress["n"] += len(chunk)
            if stats["calls"] % 10 == 0:
                el = time.time() - t0
                rate = (progress["n"] - done) / max(el, 1e-9)
                left = (n - progress["n"]) / max(rate, 1e-9)
                print(f"  [{tag}] {progress['n']:,}/{n:,} · {rate:.0f}건/초 · "
                      f"잔여 {left / 60:.0f}분", flush=True)

    starts = list(range(done, n, BATCH))
    if starts:
        from concurrent.futures import ThreadPoolExecutor

        with ThreadPoolExecutor(max_workers=workers) as ex:
            # 청크 단위로 제출하고 중간중간 진행 상태를 저장해 재개를 보장한다
            for i in range(0, len(starts), workers * 5):
                batch = starts[i : i + workers * 5]
                list(ex.map(work, batch))
                mm.flush()
                safe = min(batch[-1] + BATCH, n)
                done_path.write_text(json.dumps({"n": n, "dim": DIM, "done": safe}))
        mm.flush()
        done_path.write_text(json.dumps({"n": n, "dim": DIM, "done": n}))

    stats["uniq"][tag] = n
    return mm, inverse


def normalize_rows(mat: np.ndarray, sl: slice) -> np.ndarray:
    """잘린 벡터는 L2 norm이 1이 아니다. 정규화 없이 코사인을 재면 안 된다."""
    block = np.asarray(mat[sl], dtype=np.float32)
    norms = np.linalg.norm(block, axis=1, keepdims=True)
    return block / np.clip(norms, 1e-9, None)


def topk_for_queries(mat, qmat, k, chunk=50000):
    """고유값 행렬을 조각내 훑어 질의별 top-k를 모은다.

    30만 x 768(940MB)을 통째로 메모리에 올리지 않는다. 조각마다 후보를 뽑아
    누적하고 마지막에 정렬한다. 벡터 검색 엔진이 하는 일의 가장 단순한 형태다.
    """
    n, nq = mat.shape[0], qmat.shape[0]
    acc = [[] for _ in range(nq)]
    for s in range(0, n, chunk):
        e = min(s + chunk, n)
        sims = normalize_rows(mat, slice(s, e)) @ qmat.T  # (chunk, nq)
        for qi in range(nq):
            col = sims[:, qi]
            take = min(k, col.shape[0])
            loc = np.argpartition(-col, take - 1)[:take]
            acc[qi].extend(zip(col[loc].tolist(), (loc + s).tolist()))
    return [sorted(a, key=lambda x: -x[0])[:k] for a in acc]


def main() -> int:
    ap = argparse.ArgumentParser(description="임베딩 조립안 A/B 프로브")
    ap.add_argument("--input", default="data/sample/raw_sample.csv")
    ap.add_argument("--variants", default="A,B,C,D")
    ap.add_argument("--rows", type=int, default=0)
    ap.add_argument("--top-k", type=int, default=5)
    ap.add_argument("--workers", type=int, default=1)
    ap.add_argument("--report", default="reports/embed_probe.txt")
    args = ap.parse_args()

    which = [v.strip().upper() for v in args.variants.split(",") if v.strip()]
    df = load_frame(args.input)
    if args.rows:
        df = df.head(args.rows).reset_index(drop=True)
    variants = build_texts(df, which)

    stats = {"calls": 0, "tokens": 0, "retries": 0, "uniq": {}}
    lines = []

    def out(s=""):
        print(s, flush=True)
        lines.append(s)

    out(f"[probe] {MODEL} · {DIM}차원 · {len(df):,}건 · 조립안 {'/'.join(which)}"
        f" · 동시 {args.workers}")
    out()
    out("[조립안 실물]")
    for label, texts in variants.items():
        out(f"  {label:<18} {texts[41][:76]}")
    out()
    out("[평균 길이] " + " · ".join(
        f"{k.split()[0]} {sum(len(t) for t in v) / len(v):.1f}자" for k, v in variants.items()))
    out()

    t0 = time.time()
    mats = {}
    for label, texts in variants.items():
        tag = label.split()[0]
        print(f"  임베딩: {label}")
        mats[label] = embed_unique(texts, "RETRIEVAL_DOCUMENT", tag, args.workers, stats)
        u = stats["uniq"][tag]
        out(f"[{tag}] 고유 텍스트 {u:,}건 / {len(texts):,}행"
            f" (중복 제거 {100 * (1 - u / len(texts)):.1f}%)")

    # np.unique는 정렬한다. inverse로 되돌리지 않으면 질의와 결과가 어긋난다
    qm, qinv = embed_unique(QUERIES, "RETRIEVAL_QUERY", "Q", 1, stats)
    qmat = normalize_rows(qm, slice(0, qm.shape[0]))[qinv]
    elapsed = time.time() - t0

    # 질의별로 모든 조립안을 나란히 보여준다 — 판정은 사람이 한다
    per_variant = {}
    for label, (mat, inverse) in mats.items():
        first_row = {}
        for row_i, u_i in enumerate(inverse):
            first_row.setdefault(int(u_i), row_i)
        tops = topk_for_queries(mat, qmat, args.top_k)
        per_variant[label] = [[(sc, first_row[i]) for sc, i in t] for t in tops]

    out()
    out("=" * 78)
    for qi, q in enumerate(QUERIES):
        out(f'\n■ "{q}"')
        for label, tops in per_variant.items():
            rows = [r for _, r in tops[qi]]
            big = df.iloc[rows]["식품대분류명"].tolist()
            focus = max(big.count(c) for c in set(big)) / len(big) * 100
            out(f"  {label:<18} 최고 {tops[qi][0][0]:.3f} · 대분류 집중도 {focus:.0f}%")
            for sc, r in tops[qi][:3]:
                out(f"      {sc:.3f}  {df.iloc[r]['식품명'][:32]:<34}"
                    f" [{df.iloc[r]['식품중분류명']}]")
    out()
    out("=" * 78)
    cost = stats["tokens"] / 1_000_000 * 0.15
    out(f"[cost] 호출 {stats['calls']:,}회 · 입력 {stats['tokens']:,} tok · "
        f"약 ${cost:.2f} · 429 재시도 {stats['retries']}회")
    out(f"[time] {elapsed / 60:.1f}분")
    out("       벡터는 data/clean/.probe_*.f32에 남습니다 (재실행은 공짜, 재개 가능)")

    report = Path(args.report)
    report.parent.mkdir(parents=True, exist_ok=True)
    report.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"\n[done] → {report}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
