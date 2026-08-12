"""K-FIND 가공식품 DB 원본 받기 — 파일 다운로드 경로 (기본 권장).

원본은 레포 밖에 있고, 받는 방법은 레포 안에 있다. 이 파일이 그 "방법"이다.
187MB를 커밋하는 대신 출처·버전·검증 절차를 커밋한다. 재현성은 파일이 아니라
절차가 보장한다.

이 스크립트가 하는 일:
  1. 제공 중인 가공식품 DB 버전 목록 조회 (K-FIND 내부 JSON)
  2. 다운로드 페이지 안내 — 실제 내려받기는 브라우저에서 사람이 한다
  3. 받은 파일을 data/raw/에 놓고 검증 (크기·행수·컬럼·SHA-256)

왜 다운로드까지 자동화하지 않는가:
  K-FIND는 파일을 주기 전에 "기관유형·소속·사용목적"을 묻는 짧은 설문을 받는다.
  인증이 아니라 설문이고, 답은 사람이 해야 한다. 스크립트가 아무 값이나 채워
  통과시키는 것은 데이터 제공자와의 약속을 우회하는 일이다. 자동화하지 않는
  것도 설계다 — 대신 그 앞뒤(목록 조회, 받은 뒤 검증)를 자동화한다.

사용법:
  docker compose exec api uv run python scripts/download_file.py            # 버전 목록 + 안내
  docker compose exec api uv run python scripts/download_file.py --verify data/raw/가공식품DB.xlsx
"""

import argparse
import hashlib
import json
import sys
import urllib.request
from pathlib import Path

BASE = "https://various.foodsafetykorea.go.kr/nutrient"
LIST_API = f"{BASE}/export/down/historyJson.do"
DOWNLOAD_PAGE = f"{BASE}/general/down/list.do"
HISTORY_PAGE = f"{BASE}/general/down/historyList.do"

DB_GUBUN_PROCESSED = "01"  # 가공식품 DB (음식 DB는 3주차의 소재였다)

EXPECTED_KEY_COLUMNS = [
    "식품코드",
    "식품명",
    "식품대분류명",
    "1회 섭취참고량",
    "에너지(kcal)",
    "나트륨(mg)",
]


def fetch_versions() -> list[dict]:
    """제공 중인 가공식품 DB 버전 목록. 읽기 전용 조회라 설문과 무관하다."""
    body = f"searchDbGubun={DB_GUBUN_PROCESSED}&pagesize=20".encode()
    req = urllib.request.Request(
        LIST_API,
        data=body,
        headers={
            "X-Requested-With": "XMLHttpRequest",
            "Content-Type": "application/x-www-form-urlencoded",
        },
    )
    with urllib.request.urlopen(req, timeout=30) as res:
        payload = json.load(res)
    return payload.get("returnMap", {}).get("dataList", [])


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def verify(path: Path) -> int:
    """받은 원본이 쓸 수 있는 물건인지 확인한다.

    파이프라인을 돌리기 전에 여기서 걸러야 한다. 3시간짜리 배치가
    "컬럼이 없다"로 죽는 것보다 30초짜리 검증이 낫다.
    """
    if not path.exists():
        print(f"[verify] 파일이 없습니다: {path}", file=sys.stderr)
        return 1

    size_mb = path.stat().st_size / 1024 / 1024
    print(f"[verify] {path.name}")
    print(f"[verify] 크기: {size_mb:,.2f} MiB")

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[verify] openpyxl이 없어 내용 검증은 건너뜁니다 (크기만 확인)")
        return 0

    # read_only 스트리밍 — 187MB를 통째로 메모리에 올리지 않는다
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [("" if c is None else str(c)) for c in next(rows)]
    n = sum(1 for _ in rows)

    print(f"[verify] 시트: {wb.sheetnames[0]}")
    print(f"[verify] 행수: {n:,} · 열수: {len(header)}")

    missing = [c for c in EXPECTED_KEY_COLUMNS if c not in header]
    if missing:
        print(f"[verify] ✗ 필수 컬럼 없음: {missing}", file=sys.stderr)
        print("[verify]   DB 버전이 바뀌었을 수 있습니다 — 파이프라인의 컬럼 상수를 확인하세요")
        return 1

    print(f"[verify] ✓ 필수 컬럼 {len(EXPECTED_KEY_COLUMNS)}개 확인")
    print(f"[verify] SHA-256: {sha256(path)}")
    print("[verify]   이 해시를 README에 적어두면 '같은 원본'을 증명할 수 있습니다")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="K-FIND 가공식품 DB 원본 받기 (파일 경로)")
    parser.add_argument("--verify", metavar="PATH", help="이미 받은 원본 파일을 검증한다")
    args = parser.parse_args()

    if args.verify:
        return verify(Path(args.verify))

    print("[list] 제공 중인 가공식품 DB 버전을 조회합니다…")
    try:
        versions = fetch_versions()
    except Exception as exc:  # noqa: BLE001 — 네트워크 문제는 안내로 흡수한다
        print(f"[list] 목록 조회 실패: {exc}")
        versions = []

    if versions:
        print(f"[list] {len(versions)}건")
        for v in versions:
            print(f"  - {v.get('dbVersion')}  {v.get('fileSizeMb'):>10}  ({v.get('projectOrgan')})")
    else:
        print("[list] 목록을 가져오지 못했습니다. 아래 페이지에서 직접 확인하세요.")

    print()
    print("─" * 68)
    print("내려받기는 브라우저에서 진행합니다. K-FIND가 파일을 주기 전에")
    print("기관유형·소속·사용목적을 묻는 짧은 설문을 받기 때문입니다.")
    print()
    print(f"  1) {DOWNLOAD_PAGE}")
    print("     → [DB 다운로드 받기] → 가공식품 DB 선택 → 설문 작성 → 엑셀 저장")
    print(f"     (지난 버전은 {HISTORY_PAGE})")
    print("  2) 받은 파일을 data/raw/ 에 둡니다 (레포에는 커밋되지 않습니다)")
    print("  3) 검증:")
    print("     uv run python scripts/download_file.py --verify data/raw/<받은파일>.xlsx")
    print("─" * 68)
    print()
    print("무거운 배치를 직접 돌릴 생각이 없다면 원본은 받지 않아도 됩니다.")
    print("릴리즈 자산의 정제 산출물(parquet·덤프·인덱스)로 v1.0을 바로 띄울 수 있습니다.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
