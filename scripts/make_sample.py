"""원본 306,307건 → data/sample/raw_sample.csv (1,000건).

이 샘플이 레포에 커밋되는 유일한 데이터다. 다운로드 없이도 "치우기 전의
원본"을 관찰할 수 있어야 시연 도입부가 성립하기 때문이다.

## 층화 샘플인 이유 — 더러움은 희귀하다

원본을 전수 조사해 보면 인코딩 오염은 306,307건 중 **54건**(0.018%)뿐이다.
무작위로 1,000건을 뽑으면 기대값이 0.18건, 즉 **대체로 한 건도 안 잡힌다.**
"1,000건 훑어보니 깨끗하던데요"가 나오는 이유이고, 그 판단이 31만 건 배치를
깨뜨린다.

그래서 두 층으로 나눠 뽑는다:
  - 희귀 케이스 층 (40건): BOM·전각 물음표·줄바꿈·NFD·앞뒤 공백이 실제로 든 행
  - 계통 추출 층 (960건): 파일 전체를 균등 간격으로 훑어 분포를 보존

일부러 심었다는 사실 자체가 교훈이다. 희귀 결함은 찾아 나서야 보인다.
결측률·섭취참고량 분포 같은 "전체의 성질"은 960건 쪽이 그대로 보존한다.

사용법:
  uv run python scripts/make_sample.py --input data/raw/<원본>.xlsx
"""

import argparse
import csv
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

TARGET = 1000
DIRTY_BUDGET = 40

# 인코딩 사고의 흔적들 — 대체문자, 전각 물음표, 제어문자(줄바꿈 포함), BOM
BROKEN = re.compile("[\ufffd\uff1f]")
CONTROL = re.compile("[\u0000-\u0008\u000a-\u001f\u007f]")
BOM = "\ufeff"


def dirt_tags(name: str) -> list[str]:
    """식품명 한 칸에서 발견되는 오염 유형. 사유를 이름 붙여 남기는 것이 정제의 시작이다."""
    tags = []
    if name.startswith(BOM):
        tags.append("bom")
    if BROKEN.search(name):
        tags.append("broken_char")
    if CONTROL.search(name):
        tags.append("control_char")
    if name and unicodedata.normalize("NFC", name) != name:
        tags.append("nfd")
    if name != name.strip():
        tags.append("edge_space")
    return tags


def main() -> int:
    parser = argparse.ArgumentParser(description="원본 엑셀에서 층화 샘플 1,000건 추출")
    parser.add_argument("--input", required=True, help="K-FIND 가공식품 DB 엑셀")
    parser.add_argument("--output", default="data/sample/raw_sample.csv")
    parser.add_argument("--total-hint", type=int, default=306307, help="계통 추출 간격 계산용")
    args = parser.parse_args()

    # read_only 스트리밍 — 187MB를 통째로 메모리에 올리지 않는다
    wb = load_workbook(args.input, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [("" if c is None else str(c)) for c in next(rows)]
    ncol = len(header)
    name_idx = header.index("식품명")

    stride = max(1, args.total_hint // (TARGET - DIRTY_BUDGET))
    dirty, systematic = [], []
    n = 0

    for row in rows:
        n += 1
        cells = ["" if v is None else str(v) for v in row[:ncol]]
        cells += [""] * (ncol - len(cells))

        if dirt_tags(cells[name_idx]) and len(dirty) < DIRTY_BUDGET:
            dirty.append(cells)
        elif n % stride == 0 and len(systematic) < TARGET - DIRTY_BUDGET:
            systematic.append(cells)

        if n % 50000 == 0:
            print(f"  {n:,}행 스캔", flush=True)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(dirty + systematic)

    print(f"[sample] 원본 {n:,}행 · {ncol}열")
    print(f"[sample] 희귀 케이스 {len(dirty)}건 + 계통 추출 {len(systematic)}건 = {len(dirty) + len(systematic)}건")
    print(f"[sample] → {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
