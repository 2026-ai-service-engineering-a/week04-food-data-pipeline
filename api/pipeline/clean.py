"""1단계 정제 — 인코딩 정리, 깨진 행 분리, 컬럼 슬림, 결측 정책.

원본 306,307행 × 166열을 17열로 줄인다. **버리는 게 아니라 안 싣는 것**이고,
원본은 data/raw/에 그대로 있다.

## 슬림의 기준은 결측률이 아니다

전수 조사해 보면 결측이 무작위가 아니다. 166열이 꽉 찬 38열과 텅 빈 125열로
갈리고, 그 사이는 셋뿐이다. 라벨에 적을 의무가 있으면 채워지고, 없으면 수집될
원천 자체가 없기 때문이다.

그런데 결측 1% 미만인 열이 38개인데도 17개만 싣는다. **기준은 "이 서비스가
쓰는가"**이고, 결측률은 나머지를 안 실어도 되는 근거일 뿐이다. 코드와 이름이
쌍으로 있으면 이름만 싣고, 수입업체·유통업체·데이터생성방법은 검색에도 답변에도
쓰지 않는다.

## 결측은 NULL로 둔다

0으로 채우면 검색이 거짓말을 한다. 이 데이터에는 나트륨이 **진짜 0인 제품**
(튀김전용유 같은 유지류)이 실제로 있어서, 모름과 0을 섞으면 구분이 영영
불가능해진다.

사용법:
  uv run python -m pipeline.clean --input data/raw/원본.xlsx \\
    --output data/clean/foods_clean.parquet
"""

import argparse
import sys
import time
from pathlib import Path

import pandas as pd

from pipeline.report import write_clean_report
from pipeline.textclean import clean_text, has_broken_char

# 166열 중 남기는 17열.
#   식별·분류 6 — 소분류까지 싣는다. 3회전 임베딩이 대/중/소분류를 한 문장으로
#                 합쳐 쓰기 때문이다. 여기서 빼면 그때 원본을 다시 열어야 한다
#   라벨 의무 표시 9 — 결측이 0.0~0.03%인 유일한 영양성분들
#   환산용 2 — 1회 섭취참고량은 2단계가 숫자로 바꾸고, 식품중량은 포장 단위다
SLIM_COLUMNS = [
    "식품코드", "식품명", "식품대분류명", "식품중분류명", "식품소분류명", "제조사명",
    "에너지(kcal)", "단백질(g)", "지방(g)", "탄수화물(g)", "당류(g)", "나트륨(mg)",
    "포화지방산(g)", "트랜스지방산(g)", "콜레스테롤(mg)",
    "1회 섭취참고량", "식품중량",
]

TEXT_COLUMNS = [
    "식품명", "식품대분류명", "식품중분류명", "식품소분류명", "제조사명",
    "1회 섭취참고량", "식품중량",
]
# 식품중량은 '500g' 꼴의 문자열이라 숫자로 바꾸면 전부 NULL이 된다.
# 원문 그대로 싣고, 숫자가 필요하면 2단계 파서가 뽑는다.
NUMERIC_COLUMNS = [c for c in SLIM_COLUMNS if c.endswith((")", "(kcal)"))]

# 원본이 몇 열이었는지. 슬림 전 크기를 리포트에 남기려고 들고 있는다
RAW_COLUMN_COUNT: int | None = None

REJECT_BROKEN = "대체불가 문자 포함"
REJECT_NO_NAME = "식품명 결측"


def load_raw(path: Path) -> pd.DataFrame:
    """원본을 읽는다. 187MB xlsx는 필요한 열만 뽑아도 몇 분 걸린다."""
    if path.suffix == ".csv":
        return pd.read_csv(path, dtype=str)
    if path.suffix == ".parquet":
        return pd.read_parquet(path)

    from openpyxl import load_workbook

    print(f"[clean] {path.name} 읽는 중 (스트리밍)", flush=True)
    global RAW_COLUMN_COUNT
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [("" if c is None else str(c)) for c in next(rows)]
    missing = [c for c in SLIM_COLUMNS if c not in header]
    if missing:
        raise SystemExit(f"원본에 없는 컬럼: {missing} — DB 버전을 확인하세요")
    RAW_COLUMN_COUNT = len(header)
    idx = [header.index(c) for c in SLIM_COLUMNS]

    data = []
    for i, row in enumerate(rows, 1):
        data.append([row[j] for j in idx])
        if i % 100000 == 0:
            print(f"        {i:,}행", flush=True)
    return pd.DataFrame(data, columns=SLIM_COLUMNS)


def clean(df: pd.DataFrame, n_raw_cols: int | None = None) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """정리 → 분리 → 형변환. (남긴 것, 뺀 것, 집계)를 돌려준다."""
    n_in = len(df)
    df = df.reindex(columns=SLIM_COLUMNS)

    # ① 문자 정리. 고칠 수 있는 오염은 여기서 전부 고쳐진다
    before = df[TEXT_COLUMNS].copy()
    for col in TEXT_COLUMNS:
        df[col] = df[col].map(clean_text)
    changed = (before.fillna("") != df[TEXT_COLUMNS].fillna(""))
    fixed = int(changed.any(axis=1).sum())
    fixed_by_col = {c: int(n) for c, n in changed.sum().items() if n}

    # ② 분리. 삭제가 아니라 분류다 — 사유를 붙여 rejected로 보낸다
    name = df["식품명"].fillna("")
    reasons = pd.Series("", index=df.index)
    reasons[name.map(has_broken_char)] = REJECT_BROKEN
    reasons[name.str.strip() == ""] = REJECT_NO_NAME

    rejected = df[reasons != ""].copy()
    rejected["사유"] = reasons[reasons != ""]
    kept = df[reasons == ""].copy()

    # ③ 빈 문자열을 NULL로 되돌린다.
    #    pandas 3.0의 str dtype은 None을 ""로 바꿔 버린다. 그대로 두면 결측
    #    55,972건이 "값 있음"으로 둔갑하고, 이 단계가 지키려는 결측 정책이
    #    조용히 깨진다. **빈칸과 빈 문자열은 다르다.**
    for col in TEXT_COLUMNS:
        kept[col] = kept[col].replace("", None)

    # ④ 수치 형변환. 빈칸은 NULL로 남는다 — 0으로 채우지 않는다
    for col in NUMERIC_COLUMNS:
        if col in kept.columns:
            kept[col] = pd.to_numeric(kept[col], errors="coerce")

    stats = {
        "input_rows": n_in,
        "input_cols": n_raw_cols or len(df.columns),
        "output_rows": len(kept),
        "output_cols": len(SLIM_COLUMNS),
        "text_fixed": fixed,
        "text_fixed_by_col": fixed_by_col,
        "rejected": reasons[reasons != ""].value_counts().to_dict(),
    }
    return kept, rejected[["식품코드", "식품명", "사유"]], stats


def main() -> int:
    ap = argparse.ArgumentParser(description="가공식품 DB 정제 1단계")
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", default="data/clean/foods_clean.parquet")
    ap.add_argument("--report", default="reports/clean_report.txt")
    ap.add_argument("--rejected", default="reports/rejected.csv")
    args = ap.parse_args()

    src = Path(args.input)
    if not src.exists():
        print(f"입력이 없습니다: {src}", file=sys.stderr)
        return 1

    t0 = time.time()
    raw = load_raw(src)
    kept, rejected, stats = clean(raw, RAW_COLUMN_COUNT)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    kept.to_parquet(out, index=False)

    rej = Path(args.rejected)
    rej.parent.mkdir(parents=True, exist_ok=True)
    rejected.to_csv(rej, index=False)

    stats["elapsed_sec"] = time.time() - t0
    write_clean_report(Path(args.report), kept, stats, out, rej)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
